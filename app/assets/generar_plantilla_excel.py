"""
Genera la plantilla Excel para la evaluación organizacional de EBLET.
"""

import pandas as pd
from openpyxl import Workbook
from pathlib import Path

# =====================================================
# CREAR CARPETA assets SI NO EXISTE
# =====================================================

Path("assets").mkdir(exist_ok=True)

# =====================================================
# COLUMNAS
# =====================================================

columnas = [
    "empleado_id",
    "departamento",
    "seniority",
    "salario_anual"
]

# Preguntas q1-q72
columnas += [f"q{i}" for i in range(1, 73)]

# =====================================================
# CREAR LIBRO EXCEL
# =====================================================

wb = Workbook()

# -----------------------------
# Hoja Respuestas
# -----------------------------

ws = wb.active
ws.title = "Respuestas"

# Cabeceras
for i, columna in enumerate(columnas, start=1):
    ws.cell(row=1, column=i).value = columna

# Ejemplo de una fila
ejemplo = [
    "EMP001",
    "Desarrollo",
    "Senior",
    45000
]

# Respuestas ejemplo
ejemplo.extend([3] * 72)

for i, valor in enumerate(ejemplo, start=1):
    ws.cell(row=2, column=i).value = valor

# -----------------------------
# Hoja Instrucciones
# -----------------------------

ws2 = wb.create_sheet("Instrucciones")

texto = [
    "PLANTILLA EBLET",
    "",
    "INSTRUCCIONES",
    "",
    "• Cada fila corresponde a un empleado.",
    "• No modificar el nombre de las columnas.",
    "• q1-q72 deben contener valores entre 1 y 5.",
    "• salario_anual debe expresarse en euros brutos.",
    "",
    "Valores recomendados:",
    "",
    "seniority: Junior | Mid | Senior | Lead",
    "",
    "Una vez cumplimentado, guardar el archivo y subirlo a EBLET."
]

for fila, linea in enumerate(texto, start=1):
    ws2.cell(row=fila, column=1).value = linea

# =====================================================
# GUARDAR
# =====================================================

archivo = "assets/plantilla_empresa.xlsx"

wb.save(archivo)

print(f"✅ Plantilla creada correctamente en: {archivo}")